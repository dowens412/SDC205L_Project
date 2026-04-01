from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

# Ask the user for their student ID
student_id = input("Enter your student ID: ")

# Display the menu title
print(student_id + "'s Spreadsheet Automation Menu")

# Store menu options in a list
menu_options = [
    "1 Input Data",
    "2 View Current Data",
    "3 Generate Report"
]

print("Choose a number from the following options")

# Loop through the list and print each menu option
for option_text in menu_options:
    print(option_text)

# Store user selection
option = input()


# Converts Fahrenheit to Celsius
def convertData(value):
    return (value - 32) * 5 / 9


# Appends a new line of data to the CSV file
def insertData(file_path, data):
    try:
        with open(file_path, "a") as file:
            file.write(data + "\n")
    except:
        print("Error: unable to write to file")


# Reads and prints the contents of the CSV file
def viewData(file_path):
    try:
        with open(file_path, "r") as file:
            print("The file C:\\Users\\student\\ZooData.csv")
            print(file.read(), end="")
    except:
        print("Error: unable to read file")


# Handles user input and saves entries to the CSV
def getInput():
    entries = int(input("How many entries are you inputting?\n"))

    for i in range(entries):
        date = input("Enter a date:\n")
        temp = float(input("Enter the highest temp for the inputted date:\n"))

        # Convert temperature
        converted_value = convertData(temp)

        # Create CSV row
        data = str(date) + "," + str(temp) + "," + str(converted_value)

        try:
            insertData("ZooData.csv", data)
            print("The following data was saved at " + str(datetime.now()))
            print(data)
        except:
            print("Error: unable to save data")


# Creates an Excel file and generates a chart based on CSV data
def createChart(file_path, chart_type):
    # Ask user which data to use
    source_choice = input("Choose data source (1 for Fahrenheit, 2 for Celsius):\n")

    dates = []
    values = []

    # Read CSV and split into lists
    try:
        with open(file_path, "r") as file:
            for line in file:
                parts = line.strip().split(",")
                if len(parts) == 3:
                    dates.append(parts[0])

                    # Select correct column based on user choice
                    if source_choice == "1":
                        values.append(float(parts[1]))
                    else:
                        values.append(float(parts[2]))
    except:
        print("Error: unable to read csv file")
        return

    # Create workbook and worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Temperature Report"

    # Add headers
    sheet["A1"] = "Date"

    if source_choice == "1":
        sheet["B1"] = "Fahrenheit"
        y_axis_title = "Fahrenheit"
    else:
        sheet["B1"] = "Celsius"
        y_axis_title = "Celsius"

    # Insert data into Excel
    row_number = 2
    for i in range(len(dates)):
        sheet.cell(row=row_number, column=1, value=dates[i])
        sheet.cell(row=row_number, column=2, value=values[i])
        row_number += 1

    # Choose chart type
    if chart_type == "bar":
        chart = BarChart()
    else:
        chart = LineChart()

    # Set chart data and categories
    data = Reference(sheet, min_col=2, min_row=1, max_row=len(values) + 1)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=len(dates) + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Label axes and title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = "Date"
    chart.title = student_id + " " + str(datetime.now().date())

    # Add chart to sheet
    sheet.add_chart(chart, "D2")

    # Save Excel file
    workbook.save("final.xlsx")

    print("Report created successfully in final.xlsx")


# Prompts user for chart type and calls createChart
def generateReport(file_path):
    graph_type = input("Enter chart type (bar or line):\n").lower()

    if graph_type == "bar" or graph_type == "line":
        createChart(file_path, graph_type)
    else:
        print("Error: invalid chart type selected")


# Menu control
if option == "1":
    print("You selected " + option + " at " + str(datetime.now()))
    getInput()
elif option == "2":
    print("You selected " + option + " at " + str(datetime.now()))
    viewData("ZooData.csv")
elif option == "3":
    print("You selected " + option + " at " + str(datetime.now()))
    generateReport("ZooData.csv")
else:
    print("Error: The chosen functionality is not implemented yet")